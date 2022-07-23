# ！/usr/bin/env python
# -*- coding：utf-8 -*-
# 1. python-excel
# 以openpyxl为例
# excel写入
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = 'range names'
for row in range(1, 31):
    ws1.append(range(600))

ws2 = wb.create_sheet(title='pi')
ws2['A1'] = 3.15
print(ws2['A1'].value)

ws3 = wb.create_sheet(title='data')
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value='{0}'.format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title='my_sheet')
for i in range(1, 31):
    ws4.cell(column=1, row=i).value = 'test'

wb.save(filename=dest_filename)

# excel读取
from openpyxl import load_workbook

wb = load_workbook(filename='empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

for i in range(1, 31):
    print(sheet_ranges.cell(column=18, row=i).value)
